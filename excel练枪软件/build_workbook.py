from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


ROOT = Path(__file__).resolve().parent
DIST_DIR = ROOT / "dist"
OUTPUT_PATH = DIST_DIR / "Excel练枪软件-模板.xlsm"


def apply_column_widths(ws, start_col: str, end_col: str, width: float) -> None:
    for code in range(ord(start_col), ord(end_col) + 1):
        ws.column_dimensions[chr(code)].width = width


def style_button(ws, cell_range: str, label: str) -> None:
    ws.merge_cells(cell_range)
    cell = ws[cell_range.split(":")[0]]
    cell.value = label
    cell.font = Font(name="Microsoft YaHei", size=14, bold=True, color="FFFFFF")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill("solid", fgColor="2F80ED")
    edge = Side(style="medium", color="1F4E79")
    for row in ws[cell_range]:
        for item in row:
            item.border = Border(left=edge, right=edge, top=edge, bottom=edge)


def create_start_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("开始页")
    ws.sheet_view.showGridLines = False
    apply_column_widths(ws, "A", "Q", 4.5)
    for row in range(1, 28):
        ws.row_dimensions[row].height = 24

    ws.merge_cells("B2:O4")
    ws["B2"] = "Excel 练枪软件"
    ws["B2"].font = Font(name="Microsoft YaHei", size=24, bold=True, color="1F2937")
    ws["B2"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("C6:N10")
    ws["C6"] = (
        "玩法说明\n"
        "1. 双击下方开始按钮开始 30 秒训练\n"
        "2. 游戏页会随机显示文字人形，并给出目标部位\n"
        "3. 直接点击靶区单元格射击，命中指定部位得分\n"
        "4. 命中错误部位或空白区域会扣分"
    )
    ws["C6"].font = Font(name="Microsoft YaHei", size=12, color="374151")
    ws["C6"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    style_button(ws, "E13:K15", "双击这里开始游戏")

    ws.merge_cells("C18:N24")
    ws["C18"] = (
        "首次使用说明\n"
        "1. 打开本工作簿后，按 Alt + F11 进入 VBA 编辑器\n"
        "2. 导入 vba 文件夹中的 .bas 模块\n"
        "3. 打开 ThisWorkbook 对象，把 ThisWorkbook_Code.txt 中的内容粘贴进去\n"
        "4. 返回 Excel，启用宏后即可游玩"
    )
    ws["C18"].font = Font(name="Microsoft YaHei", size=11, color="4B5563")
    ws["C18"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


def create_game_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("游戏页")
    ws.sheet_view.showGridLines = False
    apply_column_widths(ws, "A", "Q", 4.0)
    for row in range(1, 28):
        ws.row_dimensions[row].height = 24
    for row in range(6, 23):
        ws.row_dimensions[row].height = 28

    labels = {
        "B2": "剩余时间",
        "E2": "当前目标",
        "H2": "分数",
        "K2": "总点击",
        "N2": "命中",
        "B3": "连击",
        "E3": "最高连击",
        "H3": "命中率",
        "K3": "反馈",
    }
    for cell, value in labels.items():
        ws[cell] = value
        ws[cell].font = Font(name="Microsoft YaHei", size=11, bold=True, color="111827")
        ws[cell].alignment = Alignment(horizontal="center", vertical="center")

    value_cells = ["C2", "F2", "I2", "L2", "O2", "C3", "F3", "I3"]
    for cell in value_cells:
        ws[cell] = "-"
        ws[cell].font = Font(name="Consolas", size=12, bold=True, color="1F2937")
        ws[cell].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("L3:O3")
    ws["L3"] = "等待开始"
    ws["L3"].font = Font(name="Microsoft YaHei", size=11, color="1F2937")
    ws["L3"].alignment = Alignment(horizontal="center", vertical="center")

    thin = Side(style="thin", color="D1D5DB")
    outer = Side(style="medium", color="111827")
    for row in ws["B6:Q22"]:
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(name="Consolas", size=14, bold=True, color="111827")
            cell.fill = PatternFill("solid", fgColor="F9FAFB")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in range(6, 23):
        for col in range(2, 18):
            cell = ws.cell(row=row, column=col)
            if row in (6, 22) or col in (2, 17):
                cell.border = Border(left=outer, right=outer, top=outer, bottom=outer)

    ws.merge_cells("B24:Q25")
    ws["B24"] = "点击靶区单元格射击。只有命中当前提示的指定部位才算有效命中。"
    ws["B24"].font = Font(name="Microsoft YaHei", size=11, color="4B5563")
    ws["B24"].alignment = Alignment(horizontal="center", vertical="center")


def create_result_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("结算页")
    ws.sheet_view.showGridLines = False
    apply_column_widths(ws, "A", "Q", 4.5)
    for row in range(1, 24):
        ws.row_dimensions[row].height = 24

    ws.merge_cells("C2:N4")
    ws["C2"] = "本局成绩"
    ws["C2"].font = Font(name="Microsoft YaHei", size=22, bold=True, color="111827")
    ws["C2"].alignment = Alignment(horizontal="center", vertical="center")

    stats = [
        ("D6", "总分", "H6"),
        ("D8", "总点击", "H8"),
        ("D10", "有效命中", "H10"),
        ("D12", "错误点击", "H12"),
        ("D14", "命中率", "H14"),
        ("D16", "最高连击", "H16"),
    ]
    for label_cell, label, value_cell in stats:
        ws[label_cell] = label
        ws[label_cell].font = Font(name="Microsoft YaHei", size=12, bold=True, color="374151")
        ws[value_cell] = "-"
        ws[value_cell].font = Font(name="Consolas", size=14, bold=True, color="111827")
        ws[label_cell].alignment = Alignment(horizontal="center", vertical="center")
        ws[value_cell].alignment = Alignment(horizontal="center", vertical="center")

    style_button(ws, "E19:K21", "双击这里再来一局")


def create_config_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("配置页")
    ws.sheet_view.showGridLines = False
    apply_column_widths(ws, "A", "F", 18)
    for row in range(1, 20):
        ws.row_dimensions[row].height = 24

    ws["A1"] = "参数"
    ws["B1"] = "值"
    ws["A1"].font = ws["B1"].font = Font(name="Microsoft YaHei", size=12, bold=True)

    config = [
        ("游戏时长（秒）", 30),
        ("命中指定部位得分", 10),
        ("命中错误部位扣分", -2),
        ("点击空白扣分", -1),
        ("连击阈值", 5),
        ("连击奖励", 5),
    ]
    for index, (label, value) in enumerate(config, start=2):
        ws[f"A{index}"] = label
        ws[f"B{index}"] = value
        ws[f"A{index}"].font = Font(name="Microsoft YaHei", size=11)
        ws[f"B{index}"].font = Font(name="Consolas", size=11)

    ws["A10"] = "说明"
    ws["A11"] = "本页数值可直接修改，VBA 会在每局开始时读取。"
    ws["A12"] = "如果要增加人物模板，请同时修改 modTemplates.bas。"


def build_workbook() -> Path:
    wb = Workbook()
    wb.remove(wb.active)
    create_start_sheet(wb)
    create_game_sheet(wb)
    create_result_sheet(wb)
    create_config_sheet(wb)
    wb.active = 0

    DIST_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    return OUTPUT_PATH


if __name__ == "__main__":
    path = build_workbook()
    print(path)
